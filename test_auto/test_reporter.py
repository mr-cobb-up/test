from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill


class ExcelReporter:
    def __init__(self, excel_path="test_results.xlsx"):
        self.excel_path = excel_path
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.current_row = 1
        self.setup_header()

    def setup_header(self):
        headers = ["테스트 일시", "테스트 케이스", "결과", "에러 메시지", "실행 시간(초)"]
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)

    def add_result(self, test_name, result, error_message="", duration=0):
        self.current_row += 1
        row = self.current_row

        # 결과에 따른 색상 설정
        if result == "PASS":
            fill_color = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        else:
            fill_color = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # 데이터 입력
        self.ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.ws.cell(row=row, column=2, value=test_name)
        result_cell = self.ws.cell(row=row, column=3, value=result)
        result_cell.fill = fill_color
        self.ws.cell(row=row, column=4, value=error_message)
        self.ws.cell(row=row, column=5, value=round(duration, 2))

    def save(self):
        self.wb.save(self.excel_path)